#!/usr/bin/env python3
"""
Show predicted returns for ALL factors for a specific month.
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
import numpy as np
import torch
import logging
from src.data import load_data, create_rolling_windows
from src.model import SimpleNN
from src.config import get_device

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def show_all_predictions(window_index=119):  # 2015-01-01
    """
    Show predicted returns for ALL 83 factors for January 2015.
    """
    logger.info(f"Showing ALL factor predictions for window {window_index}")
    
    # Load data
    forecast_df, actual_df, factor_names, dates = load_data()
    rolling_windows = create_rolling_windows(forecast_df, actual_df, window_size=60)
    
    # Get the window data
    train_X, train_y, target_X, target_y, target_date = rolling_windows[window_index]
    
    logger.info(f"Target month: {target_date}")
    
    # Configuration
    config = {
        'hidden_sizes': [512, 256],
        'learning_rate': 0.001,
        'dropout_rate': 0.4,
        'batch_size': 32,
        'weight_decay': 1e-5,
        'n_epochs': 100,
        'early_stopping_patience': 10,
        'random_seed': 42
    }
    
    # Set random seeds for reproducibility
    import random
    torch.manual_seed(config['random_seed'])
    np.random.seed(config['random_seed'])
    random.seed(config['random_seed'])
    if torch.backends.mps.is_available():
        torch.mps.manual_seed(config['random_seed'])
    
    # Setup device and model
    device = get_device()
    n_factors = train_X.shape[1]
    model = SimpleNN(
        input_size=n_factors,
        hidden_sizes=config['hidden_sizes'],
        dropout_rate=config['dropout_rate']
    ).to(device)
    
    optimizer = torch.optim.Adam(model.parameters(), 
                                lr=config['learning_rate'], 
                                weight_decay=config['weight_decay'])
    
    # Convert to tensors
    train_X_tensor = torch.FloatTensor(train_X).to(device)
    train_y_tensor = torch.FloatTensor(train_y).to(device)
    target_X_tensor = torch.FloatTensor(target_X).to(device)
    target_y_tensor = torch.FloatTensor(target_y).to(device)
    
    # Custom top-5 loss function
    def top5_return_loss(predictions, actual_returns):
        batch_size = predictions.shape[0]
        total_loss = 0.0
        
        for i in range(batch_size):
            pred_i = predictions[i]
            actual_i = actual_returns[i]
            
            # Get top 5 predictions using temperature-scaled softmax
            temperature = 0.1
            top5_probs = torch.softmax(pred_i / temperature, dim=0)
            
            # Calculate weighted return (expectation)
            weighted_return = torch.sum(top5_probs * actual_i)
            
            # Loss is negative return (we want to maximize return)
            loss_i = -weighted_return
            total_loss += loss_i
        
        return total_loss / batch_size
    
    # Train the model (simplified training for this demo)
    logger.info("Training model...")
    
    from torch.utils.data import DataLoader, TensorDataset
    dataset = TensorDataset(train_X_tensor, train_y_tensor)
    dataloader = DataLoader(dataset, batch_size=config['batch_size'], shuffle=True)
    
    model.train()
    best_loss = float('inf')
    patience_counter = 0
    
    for epoch in range(config['n_epochs']):
        epoch_loss = 0.0
        batch_count = 0
        
        for batch_X, batch_y in dataloader:
            optimizer.zero_grad()
            predictions = model(batch_X)
            loss = top5_return_loss(predictions, batch_y)
            loss.backward()
            optimizer.step()
            epoch_loss += loss.item()
            batch_count += 1
        
        avg_epoch_loss = epoch_loss / batch_count
        
        if avg_epoch_loss < best_loss:
            best_loss = avg_epoch_loss
            patience_counter = 0
            best_model_state = model.state_dict().copy()
        else:
            patience_counter += 1
        
        if patience_counter >= config['early_stopping_patience']:
            break
    
    # Restore best model and make predictions
    model.load_state_dict(best_model_state)
    model.eval()
    
    with torch.no_grad():
        predictions = model(target_X_tensor)
        pred_values = predictions[0].cpu().numpy()
        actual_values = target_y_tensor[0].cpu().numpy()
    
    # Create comprehensive results DataFrame
    results_df = pd.DataFrame({
        'Factor': factor_names,
        'Predicted_Return': pred_values,
        'Actual_Return': actual_values,
        'Prediction_Error': pred_values - actual_values,
        'Abs_Error': np.abs(pred_values - actual_values)
    })
    
    # Add rankings
    results_df['Predicted_Rank'] = results_df['Predicted_Return'].rank(ascending=False, method='min').astype(int)
    results_df['Actual_Rank'] = results_df['Actual_Return'].rank(ascending=False, method='min').astype(int)
    results_df['Rank_Diff'] = results_df['Predicted_Rank'] - results_df['Actual_Rank']
    
    # Keep original order (as factors were read in)
    # results_df = results_df.sort_values('Predicted_Return', ascending=False).reset_index(drop=True)
    
    logger.info("="*120)
    logger.info(f"ALL FACTOR PREDICTIONS FOR {target_date}")
    logger.info("="*120)
    
    # Show top 10 predicted
    logger.info("TOP 10 PREDICTED FACTORS:")
    logger.info(f"{'Rank':<4} {'Factor':<25} {'Pred%':<8} {'Act%':<8} {'Error':<8} {'ActRank':<8}")
    logger.info("-" * 120)
    
    for i in range(10):
        row = results_df.iloc[i]
        logger.info(f"{i+1:<4} {row['Factor']:<25} {row['Predicted_Return']:<8.3f} {row['Actual_Return']:<8.3f} "
                   f"{row['Prediction_Error']:<8.3f} {row['Actual_Rank']:<8}")
    
    logger.info("\nBOTTOM 10 PREDICTED FACTORS:")
    for i in range(-10, 0):
        row = results_df.iloc[i]
        rank = len(results_df) + i + 1
        logger.info(f"{rank:<4} {row['Factor']:<25} {row['Predicted_Return']:<8.3f} {row['Actual_Return']:<8.3f} "
                   f"{row['Prediction_Error']:<8.3f} {row['Actual_Rank']:<8}")
    
    # Show actual top 10 performers
    actual_top10 = results_df.sort_values('Actual_Return', ascending=False).head(10)
    logger.info("\nACTUAL TOP 10 PERFORMERS:")
    logger.info(f"{'ActRank':<8} {'Factor':<25} {'Act%':<8} {'Pred%':<8} {'PredRank':<8}")
    logger.info("-" * 120)
    
    for i, (_, row) in enumerate(actual_top10.iterrows()):
        logger.info(f"{i+1:<8} {row['Factor']:<25} {row['Actual_Return']:<8.3f} {row['Predicted_Return']:<8.3f} "
                   f"{row['Predicted_Rank']:<8}")
    
    # Summary statistics
    logger.info("\n" + "="*120)
    logger.info("PREDICTION ACCURACY SUMMARY")
    logger.info("="*120)
    
    # Calculate correlations
    pred_actual_corr = np.corrcoef(results_df['Predicted_Return'], results_df['Actual_Return'])[0,1]
    rank_corr = np.corrcoef(results_df['Predicted_Rank'], results_df['Actual_Rank'])[0,1]
    
    # Top 5 analysis
    top5_predicted = results_df.head(5)
    top5_actual = results_df.sort_values('Actual_Return', ascending=False).head(5)
    
    overlap = len(set(top5_predicted['Factor']) & set(top5_actual['Factor']))
    
    logger.info(f"Prediction-Actual Return Correlation: {pred_actual_corr:.3f}")
    logger.info(f"Predicted-Actual Rank Correlation: {rank_corr:.3f}")
    logger.info(f"Mean Absolute Error: {results_df['Abs_Error'].mean():.3f}")
    logger.info(f"Top 5 Overlap: {overlap}/5 ({overlap/5*100:.1f}%)")
    logger.info(f"Average Predicted Return: {results_df['Predicted_Return'].mean():.3f}")
    logger.info(f"Average Actual Return: {results_df['Actual_Return'].mean():.3f}")
    
    # Save detailed results to both CSV and Excel
    import os
    os.makedirs('outputs', exist_ok=True)
    
    # Save to CSV
    csv_filename = f'outputs/all_predictions_{target_date.strftime("%Y%m")}.csv'
    results_df.to_csv(csv_filename, index=False)
    logger.info(f"\nDetailed results saved to {csv_filename}")
    
    # Save to Excel with formatting
    excel_filename = f'outputs/all_predictions_{target_date.strftime("%Y%m")}.xlsx'
    
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Main results sheet
        results_df.to_excel(writer, sheet_name='All_Predictions', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['All_Predictions']
        
        # Format headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format number columns to 3 decimal places
        from openpyxl.styles import NamedStyle
        number_style = NamedStyle(name="number_style", number_format="0.000")
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row[1:5]:  # Predicted_Return, Actual_Return, Prediction_Error, Abs_Error columns
                cell.style = number_style
        
        # Create summary sheet
        summary_data = {
            'Metric': [
                'Target Month',
                'Prediction-Actual Return Correlation',
                'Predicted-Actual Rank Correlation', 
                'Mean Absolute Error',
                'Top 5 Overlap',
                'Average Predicted Return',
                'Average Actual Return',
                'Best Predicted Factor',
                'Best Actual Factor',
                'Worst Predicted Factor',
                'Worst Actual Factor'
            ],
            'Value': [
                target_date.strftime('%Y-%m-%d'),
                f"{pred_actual_corr:.3f}",
                f"{rank_corr:.3f}",
                f"{results_df['Abs_Error'].mean():.3f}",
                f"{overlap}/5 ({overlap/5*100:.1f}%)",
                f"{results_df['Predicted_Return'].mean():.3f}",
                f"{results_df['Actual_Return'].mean():.3f}",
                results_df.iloc[0]['Factor'],
                results_df.sort_values('Actual_Return', ascending=False).iloc[0]['Factor'],
                results_df.iloc[-1]['Factor'],
                results_df.sort_values('Actual_Return', ascending=True).iloc[0]['Factor']
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        summary_worksheet = writer.sheets['Summary']
        for cell in summary_worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust summary column widths
        for column in summary_worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            summary_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create top/bottom performers sheet
        top_performers = results_df.sort_values('Actual_Return', ascending=False).head(20)
        top_performers.to_excel(writer, sheet_name='Top_20_Actual', index=False)
        
        # Format top performers sheet
        top_worksheet = writer.sheets['Top_20_Actual']
        for cell in top_worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
    logger.info(f"Excel file saved to {excel_filename}")
    logger.info("Excel file contains 3 sheets: All_Predictions, Summary, Top_20_Actual")
    
    return results_df

if __name__ == "__main__":
    results = show_all_predictions(window_index=119)  # 2015-01-01